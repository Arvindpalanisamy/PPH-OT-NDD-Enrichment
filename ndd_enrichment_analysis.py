#!/usr/bin/env python3
"""
NDD Risk Gene Enrichment Analysis
==================================
Enrichment of neurodevelopmental disorder (NDD) risk genes among
differentially expressed genes (DEGs) from mPFC and hypothalamus
RNA-seq data following postnatal oxytocin exposure.

Databases:
  - SFARI Gene (January 2026 release; confidence scores 1-3)
  - Developmental Brain Disorders (DBD)
  - SysNDD (Definitive tier; autosomal dominant subset analysed separately)

Statistical method:
  Fisher's exact test (one-tailed), comparing proportion of risk genes
  among downregulated vs upregulated DEGs (upregulated as comparator).

Usage:
  python ndd_enrichment_analysis.py

Inputs (expected in ./data/ directory):
  DEG files (Excel .xlsx):
    mPFC:
      mPFC_Oxt_v_Con_FDR_up.xlsx
      mPFC_Oxt_v_Con_FDR_down.xlsx
      mPFC_male_FDR_up.xlsx
      mPFC_male_FDR_down.xlsx
      mPFC_female_FDR_up.xlsx
      mPFC_female_FDR_down.xlsx
      mPFC_Oxt-Con_Male_vs__Female_up.xlsx   (interaction: higher in males)
      mPFC_Oxt-Con_Male_vs__Female_down.xlsx (interaction: lower in males)
    Hypothalamus:
      Hypothal_Oxt_Con_up.xlsx
      Hypothal_Oxt_Con_down.xlsx
      Hypothal_male_FDR_up.xlsx
      Hypothal_male_FDR_down.xlsx
      Hypothal_female_FDR_up.xlsx
      Hypothal_female_FDR_down.xlsx

  Database files:
    SFARI-Gene_genes_01-14-2026release_03-05-2026export.csv
    DBD-Genes-Full-Data.csv
    sysndd_panels.xlsx

Outputs:
  ndd_enrichment_mPFC.csv
  ndd_enrichment_hypothalamus.csv
  ndd_enrichment_mPFC.xlsx
  ndd_enrichment_hypothalamus.xlsx
  named_hits_mPFC.txt
  named_hits_hypothalamus.txt

Author: [Author name]
Date:   [Date]
Python: 3.x
Dependencies: openpyxl, scipy, pandas
"""

import os
import csv
import sys
import openpyxl
import pandas as pd
from scipy.stats import fisher_exact

# ── Configuration ─────────────────────────────────────────────────────────────

DATA_DIR   = "./data"
OUTPUT_DIR = "./output"

# DEG file definitions: (label, up_file, down_file)
MPFC_CONTRASTS = [
    ("Collapsed",    "mPFC_Oxt_v_Con_FDR_up.xlsx",              "mPFC_Oxt_v_Con_FDR_down.xlsx"),
    ("Male",         "mPFC_male_FDR_up.xlsx",                   "mPFC_male_FDR_down.xlsx"),
    ("Female",       "mPFC_female_FDR_up.xlsx",                 "mPFC_female_FDR_down.xlsx"),
    ("Interaction",  "mPFC_Oxt-Con_Male_vs__Female_up.xlsx",    "mPFC_Oxt-Con_Male_vs__Female_down.xlsx"),
]

HYPO_CONTRASTS = [
    ("Collapsed",    "Hypothal_Oxt_Con_up.xlsx",                "Hypothal_Oxt_Con_down.xlsx"),
    ("Male",         "Hypothal_male_FDR_up.xlsx",               "Hypothal_male_FDR_down.xlsx"),
    ("Female",       "Hypothal_female_FDR_up.xlsx",             "Hypothal_female_FDR_down.xlsx"),
]

# ── Helper functions ──────────────────────────────────────────────────────────

def load_deg_file(filepath):
    """
    Load a DEG Excel file and return a dict of {GENE_SYMBOL_UPPER: linearFC}.

    Expects columns containing 'external' (gene symbol) and 'linearfc'
    (fold change) in the header row (case-insensitive substring match).
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Empty file: {filepath}")

    headers = [str(h).lower() if h else "" for h in rows[0]]
    gene_col = next((i for i, h in enumerate(headers) if "external" in h), None)
    fc_col   = next((i for i, h in enumerate(headers) if "linearfc" in h), None)

    if gene_col is None:
        raise ValueError(f"Cannot find gene symbol column in {filepath}. "
                         f"Headers: {rows[0]}")

    genes = {}
    for row in rows[1:]:
        if row[gene_col]:
            symbol = str(row[gene_col]).strip()
            fc     = row[fc_col] if fc_col is not None else None
            genes[symbol.upper()] = {"name": symbol, "fc": fc}
    return genes


def load_sfari(filepath):
    """
    Load SFARI Gene CSV. Returns:
      sfari_all  : set of all scored gene symbols (scores 1-3)
      sfari_s1   : set of score-1 genes
      sfari_s2   : set of score-2 genes
      sfari_s3   : set of score-3 genes
    """
    sfari_all, s1, s2, s3 = set(), set(), set(), set()
    with open(filepath, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            g     = row["gene-symbol"].strip().upper()
            score = row["gene-score"].strip()
            if score in ("1", "2", "3"):
                sfari_all.add(g)
            if score == "1": s1.add(g)
            if score == "2": s2.add(g)
            if score == "3": s3.add(g)
    return sfari_all, s1, s2, s3


def load_dbd(filepath):
    """
    Load DBD CSV. Returns dict of disorder category -> set of gene symbols.
    Columns used: Gene, Autism, Epilepsy, IntellectualDisability (flexible match).
    """
    dbd_genes = {}
    with open(filepath, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            g = row["Gene"].strip().upper()
            for col, key in [("Autism", "autism"), ("Epilepsy", "epilepsy")]:
                val = row.get(col, "").strip()
                if val not in ("", "0", "No"):
                    dbd_genes.setdefault(key, set()).add(g)
    return dbd_genes


def load_sysndd(filepath):
    """
    Load SysNDD Excel file (Definitive tier only, as exported).
    Returns:
      sysndd_def  : all Definitive genes
      sysndd_ad   : Definitive + autosomal dominant inheritance
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb["data"]
    rows = list(ws.iter_rows(values_only=True))[1:]  # skip header

    sysndd_def, sysndd_ad = set(), set()
    for row in rows:
        category    = str(row[1]).strip() if row[1] else ""
        inheritance = str(row[2]).strip() if row[2] else ""
        symbol      = str(row[3]).strip().upper() if row[3] else ""
        if not symbol:
            continue
        if category == "Definitive":
            sysndd_def.add(symbol)
            if "dominant" in inheritance.lower():
                sysndd_ad.add(symbol)
    return sysndd_def, sysndd_ad


def fisher_enrichment(down_genes, up_genes, reference_set):
    """
    One-tailed Fisher's exact test: are downregulated genes more enriched
    for reference_set genes than upregulated genes?

    Contingency table:
                  In reference   Not in reference
      Down DEGs:      a               b
      Up DEGs:        c               d

    Returns: (n_down_hits, n_up_hits, odds_ratio, p_value)
    """
    a = len(set(down_genes) & reference_set)
    b = len(down_genes) - a
    c = len(set(up_genes)  & reference_set)
    d = len(up_genes) - c
    odds_ratio, p_value = fisher_exact([[a, b], [c, d]], alternative="greater")
    return a, c, round(odds_ratio, 2), p_value


def format_p(p):
    """Format p-value for output."""
    if p >= 0.05:
        return "n.s."
    return f"{p:.2e}"


def run_enrichment(contrasts, region_label, databases, output_rows, named_hits):
    """
    Run enrichment for a set of contrasts and append results to output_rows.
    Named gene hits for significant results appended to named_hits.
    """
    for contrast_label, up_file, down_file in contrasts:
        up_path   = os.path.join(DATA_DIR, up_file)
        down_path = os.path.join(DATA_DIR, down_file)

        if not os.path.exists(up_path):
            print(f"  WARNING: missing {up_path}, skipping.")
            continue
        if not os.path.exists(down_path):
            print(f"  WARNING: missing {down_path}, skipping.")
            continue

        up   = load_deg_file(up_path)
        down = load_deg_file(down_path)

        print(f"  {region_label} | {contrast_label}: "
              f"{len(down)} down, {len(up)} up")

        for db_label, ref_set in databases:
            n_dn, n_up, OR, p = fisher_enrichment(down, up, ref_set)
            sig = "SIG" if p < 0.05 else "n.s."

            row = {
                "Region":           region_label,
                "Contrast":         contrast_label,
                "Database":         db_label,
                "N_down_DEGs":      len(down),
                "N_up_DEGs":        len(up),
                "N_down_hits":      n_dn,
                "Pct_down_hits":    round(100 * n_dn / len(down), 1) if down else 0,
                "N_up_hits":        n_up,
                "Pct_up_hits":      round(100 * n_up / len(up), 1) if up else 0,
                "Odds_ratio":       OR,
                "P_value":          p,
                "Significance":     sig,
            }
            output_rows.append(row)

            # Collect named hits for significant results
            if p < 0.05 and n_dn > 0:
                hits = sorted([down[g]["name"] for g in set(down) & ref_set])
                named_hits.append({
                    "Region":    region_label,
                    "Contrast":  contrast_label,
                    "Database":  db_label,
                    "Direction": "Downregulated",
                    "N_hits":    n_dn,
                    "Genes":     ", ".join(hits),
                })


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # ── Load databases ────────────────────────────────────────────────────────
    print("Loading reference databases...")

    sfari_path  = os.path.join(DATA_DIR, "SFARI-Gene_genes_01-14-2026release_03-05-2026export.csv")
    dbd_path    = os.path.join(DATA_DIR, "DBD-Genes-Full-Data.csv")
    sysndd_path = os.path.join(DATA_DIR, "sysndd_panels.xlsx")

    for p in [sfari_path, dbd_path, sysndd_path]:
        if not os.path.exists(p):
            sys.exit(f"ERROR: Required database file not found: {p}")

    sfari_all, s1, s2, s3 = load_sfari(sfari_path)
    dbd_genes              = load_dbd(dbd_path)
    sysndd_def, sysndd_ad  = load_sysndd(sysndd_path)

    print(f"  SFARI: {len(sfari_all)} genes total "
          f"(S1={len(s1)}, S2={len(s2)}, S3={len(s3)})")
    print(f"  DBD autism: {len(dbd_genes.get('autism', set()))} genes")
    print(f"  SysNDD Definitive: {len(sysndd_def)} genes "
          f"(AD subset: {len(sysndd_ad)})")

    databases = [
        ("SFARI_All",       sfari_all),
        ("SFARI_Score1",    s1),
        ("SFARI_Score2",    s2),
        ("SFARI_Score3",    s3),
        ("DBD_Autism",      dbd_genes.get("autism", set())),
        ("DBD_Epilepsy",    dbd_genes.get("epilepsy", set())),
        ("SysNDD_Definitive",    sysndd_def),
        ("SysNDD_Definitive_AD", sysndd_ad),
    ]

    # ── Run enrichment ────────────────────────────────────────────────────────
    all_rows   = []
    named_hits = []

    print("\nmPFC enrichment analysis...")
    run_enrichment(MPFC_CONTRASTS, "mPFC", databases, all_rows, named_hits)

    print("\nHypothalamus enrichment analysis...")
    run_enrichment(HYPO_CONTRASTS, "Hypothalamus", databases, all_rows, named_hits)

    # ── Write outputs ─────────────────────────────────────────────────────────
    df = pd.DataFrame(all_rows)

    for region in ["mPFC", "Hypothalamus"]:
        sub = df[df["Region"] == region]
        csv_path  = os.path.join(OUTPUT_DIR, f"ndd_enrichment_{region}.csv")
        xlsx_path = os.path.join(OUTPUT_DIR, f"ndd_enrichment_{region}.xlsx")
        sub.to_csv(csv_path,  index=False)
        sub.to_excel(xlsx_path, index=False)
        print(f"\nSaved: {csv_path}")
        print(f"Saved: {xlsx_path}")

    # Named hits
    hits_path = os.path.join(OUTPUT_DIR, "named_hits.txt")
    with open(hits_path, "w") as f:
        f.write("NDD Risk Gene Hits (significant enrichments, downregulated DEGs)\n")
        f.write("=" * 70 + "\n\n")
        for h in named_hits:
            f.write(f"Region:   {h['Region']}\n")
            f.write(f"Contrast: {h['Contrast']}\n")
            f.write(f"Database: {h['Database']}\n")
            f.write(f"N hits:   {h['N_hits']}\n")
            f.write(f"Genes:    {h['Genes']}\n")
            f.write("-" * 70 + "\n")
    print(f"Saved: {hits_path}")

    # Print summary to console
    print("\n" + "=" * 70)
    print("SUMMARY — Significant enrichments (p < 0.05)")
    print("=" * 70)
    sig = df[df["P_value"] < 0.05][
        ["Region","Contrast","Database","N_down_hits","N_down_DEGs",
         "Odds_ratio","P_value"]
    ]
    print(sig.to_string(index=False))
    print("\nDone.")


if __name__ == "__main__":
    main()
